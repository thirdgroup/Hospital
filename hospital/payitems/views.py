from django.shortcuts import render, HttpResponse
from rest_framework import viewsets
from rest_framework import permissions
from rest_framework.request import Request
from rest_framework.response import Response
from payitems.serializers import PayitemsSerializer
from payitems.serializers import RegisterItemsSerializer
from payitems.serializers import RegistrationSerializer
from payitems.serializers import StandartResultsSetPagination
from payitems.permissions import IsPyitemsOrReadOnly, IsRegisterItemsOrReadOnly
from database.models import PayItems, RegisterItems, User, Registration
from django.contrib.auth import authenticate, login, logout
from django.views import View
from django.utils.deprecation import MiddlewareMixin

# Create your views here.

class PayItemsView(viewsets.ModelViewSet):
    """
        项目收费
    """
    pagination_class = StandartResultsSetPagination     # 分页
    # 权限认证
    permission_classes = (permissions.IsAuthenticated, IsPyitemsOrReadOnly)

    queryset = PayItems.objects.all()
    serializer_class = PayitemsSerializer

    # 对项目名称的模糊查询
    def list(self, request, *args, **kwargs):
        keyword = request.GET.get('item_name')  # 获取参数
        page = self.paginate_queryset(self.queryset)

        if page is not None:
            serializer = PayitemsSerializer(page, many=True, context={'request': self.request})
            return self.get_paginated_response(serializer.data)
        if keyword is not None:  # 如果参数不为空
            # 执行filter()方法
            queryset = PayItems.objects.filter(item_name__contains=keyword)
        else:
            # 如果参数为空，执行all()方法
            queryset = PayItems.objects.all()
        serializer = PayitemsSerializer(queryset, many=True, context={'request': self.request})
        return Response(serializer.data)  # 最后返回经过序列化的数据



class RegisterItemsView(viewsets.ModelViewSet):
    """
        项目收费登记增删改接口
    """
    pagination_class = StandartResultsSetPagination  # 分页
    # 权限认证
    permission_classes = (permissions.IsAuthenticated, IsRegisterItemsOrReadOnly)

    queryset = RegisterItems.objects.all()
    serializer_class = RegisterItemsSerializer


class RegistrationView(viewsets.ModelViewSet):
    """
        项目收费登记展示接口
    """
    pagination_class = StandartResultsSetPagination  # 分页
    # 权限认证
    permission_classes = (permissions.IsAuthenticated, IsRegisterItemsOrReadOnly)

    queryset = Registration.objects.all()
    serializer_class = RegistrationSerializer

    # 对病历号和姓名的模糊查询
    def list(self, request, *args, **kwargs):
        queryset = self.queryset
        sick_id = request.GET.get('sick_id', None)  # 获取病历号
        if sick_id is not None:
            sick_id_all_list = list(Registration.objects.values_list('id', flat=True).order_by('id'))
            if sick_id_all_list:
                sick_id_list = [i for i in sick_id_all_list if sick_id in str(i)]
                queryset = queryset.filter(id__in=sick_id_list)

        sick_name = request.GET.get('sick_name', None)    # 获取姓名
        print('sick_name',sick_name)
        if sick_name is not None:  # 如果参数不为空
            # 执行filter()方法
            queryset = queryset.filter(name__contains=sick_name)
            
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = RegistrationSerializer(page, many=True, context={'request': self.request})
            return self.get_paginated_response(serializer.data)

        serializer = RegistrationSerializer(queryset, many=True, context={'request': self.request})
        return Response(serializer.data)  # 最后返回经过序列化的数据


# class UserViewSet(viewsets.ModelViewSet):
#     """
#         用户的增删改查
#     """
#     queryset = User.objects.all()
#     serializer_class = UserSerializer




class DisableCSRFCheck(MiddlewareMixin):
    """
        强制取消CSRF
    """
    def process_request(self, request):
        setattr(request, '_dont_enforce_csrf_checks', True)


'''
def export_excel(request):
    """
        调用此接口，导出此ID患者的收费项目登记表
    """
    import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from io import BytesIO
    from django.utils.http import urlquote
    wb = Workbook()             # 新建一个 excel 文件
    wb.encoding = 'utf-8'       
    sheet1 = wb.active          # 获取第一个工作表（sheet1）
    sheet1.title = '用户收费项目登记'   # 给工作表1设置标题
    row_one = []
    for i in range(1, len(row_one)+1):
        sheet1.cell(row=1, column=i).value=row_one[i-1]
    all_obj = []
    for obj in all_obj:
        max-row = sheet1.max_row + 1 
        obj_info = [obj.id,...]
        for x in range(1, len(obj_info)+1):
            sheet1.cell(row=max_row, column=x).value = obj_info[x-1]
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '%s收费项目登记%s.xls'%(username, ctime)
    file_name = urlquote(file_name)
    response['Content-Disposition'] = 'attachment; filename=%s'%file_name
    return response
'''

def pay_items_count(request):
    """
        调用此接口，返回数据库收费项目的数量加1的结果（为 post 时展示编号）
    """
    count = PayItems.objects.count()
    return HttpResponse(count+1)


class UserAuthenticate(View):
    """
        自己写的登录接口
    """

    def post(self, request):
        # 登录
        username = request.POST.get('username')
        password = request.POST.get('password')
        print(username, password)
        users = User.objects.get(username=username)
        check = users.check_password(password)
        print('check', check)
        user = authenticate(username=username, password=password)
        print(user)
        if user:
            login(request, user)
            return HttpResponse('ok')
        else:
            return HttpResponse('错误了')

    def put(self, request):
        # 登出
        logout(request)
        return HttpResponse('退出了')

