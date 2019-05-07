import datetime

# from django.shortcuts import render
from database.models import Registration, Department, DoctorManage
from outpatient.serializers import RegistrationSerializer, RegistSerializer, DepartmentSerializer, \
    DoctorManageSerializer
from rest_framework import viewsets
from rest_framework import permissions
from rest_framework.filters import OrderingFilter, SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination, CursorPagination
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
# from rest_framework.parsers import JSONParser
from django.http import HttpResponse
from collections import OrderedDict
from rest_framework.utils.urls import replace_query_param, remove_query_param


# Create your views here.
class MyCursorPagination(CursorPagination):
    cursor_query_param = 'cursor'
    page_size = 2
    page_size_query_param = 'size'
    max_page_size = 100
    ordering = 'id'


class StandardPageNumberPagination(PageNumberPagination):
    """
    配置分页规则
    """
    page_size = 1
    page_size_query_param = 'page_size'
    page_query_param = 'page'
    max_page_size = 100

    def first_page(self):  # 自定义首页
        url = self.request.build_absolute_uri()
        return remove_query_param(url, self.page_query_param)

    def last_page(self):  # 自定义最后一页
        url = self.request.build_absolute_uri()
        count_num = self.page.paginator.count
        return replace_query_param(url, self.page_query_param, count_num)

    def get_paginated_response(self, data):
        return Response(OrderedDict([
            ('count', self.page.paginator.count),
            ('first', self.first_page()),
            ('previous', self.get_previous_link()),
            ('next', self.get_next_link()),
            ('last', self.last_page()),
            ('results', data)
        ]))


class DepartmentViewSet(viewsets.ModelViewSet):
    """
    科室信息
    """
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = (permissions.IsAuthenticated,)


# @method_decorator(csrf_exempt, name='patch_all')
class RegistrationViewSet(viewsets.ModelViewSet):
    """
    患者挂号信息
    """
    queryset = Registration.objects.all()
    serializer_class = RegistrationSerializer
    permission_classes = (permissions.IsAuthenticated,)
    pagination_class = StandardPageNumberPagination

    # filter_backends = (DjangoFilterBackend, OrderingFilter, SearchFilter,)  # 过滤后端（包括Django过滤后端、排序过滤、查询）
    # filter_fields = ('id', 'doctor', 'department')  # 过滤的字段
    # search_fields = ('id', 'doctor__real_name', 'department__department_name')  # 查询的字段
    # ordering_fields = ('id', 'regist_date',)  # 排序的字段
    # ordering = ('regist_date',)  # 按照挂号时间进行排序

    def list(self, request, *args, **kwargs):
        """
        自定义 'list' 方法，用来覆盖DRF的 'list' 方法，使用自定义的第二个序列化器（只序列化部分字段）
        """
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = RegistSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)

        serializer = RegistSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

    def get_queryset(self):
        """
        自定义get_queryset方法，用来条件查询queryset
        :return:
        """
        query_set = self.queryset

        record_no = self.request.query_params.get('record_no', None)
        if record_no:  # 对于病历号的模糊查询
            id_list = Registration.objects.values_list('id')
            id_list = list(id_list)
            new_list = list()
            if id_list:
                new_list = [str(i[0]) for i in id_list if record_no in str(i[0])]
            query_set = Registration.objects.filter(id__in=new_list)

        doctor = self.request.query_params.get('doctor', None)
        if doctor:  # 对于主治医生的模糊查询
            query_set = query_set.filter(doctor__real_name__icontains=doctor)

        department = self.request.query_params.get('department', None)
        if department:  # 对于门诊科室的模糊查询
            query_set = query_set.filter(department__department_name__icontains=department)

        start = self.request.query_params.get('start_time', None)
        end = self.request.query_params.get('end_time', None)
        if start and end:  # 对于挂号时间的模糊查询
            query_set = query_set.filter(regist_date__gte=start).filter(regist_date__lte=end)
        return query_set

    def get_obj(self, pk):
        try:
            return Registration.objects.get(pk=pk)
        except Registration.DoesNotExist:
            return None

    @action(detail=False, methods=['GET', 'POST'])
    def patch_all(self, request, *args, **kwargs):
        """
        更改所有传入的挂号ID的状态字段为：已退号；
        """
        patch_info = request.data
        print(patch_info)
        record_no_str = patch_info.get('record_no_list', None)
        if record_no_str:
            try:
                record_no_list = eval(record_no_str)
            except (TypeError, ValueError, SyntaxError):
                record_no_list = None
        else:
            record_no_list = None
        if record_no_list:
            new_list = []
            for i in record_no_list:
                obj = self.get_obj(i)
                if obj:
                    if obj.status == 5:
                        obj.status = 6
                        obj.save()
                    new_list.append(obj)
            serializer = RegistrationSerializer(new_list, many=True, context={'request': request})
            return Response(serializer.data)
        return Response({'detail': '请传递正确的参数。'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET', 'POST'])
    def export_excel(self, request, *args, **kwargs):
        """
        导出所有传入ID的挂号信息，使用Excel；
        """
        export_info = request.data
        record_no_str = export_info.get('record_no_list', None)
        if record_no_str:
            try:
                record_no_list = eval(record_no_str)
            except (TypeError, ValueError, SyntaxError):
                record_no_list = None
        else:
            record_no_list = None
        if record_no_list:
            from openpyxl import Workbook
            from io import BytesIO  # 将文件内容读写到内存中
            from django.utils.http import urlquote  # 用来解决Excel文件中文文件名无法使用问题
            wb = Workbook()
            wb.encoding = 'utf-8'
            sheet = wb.active
            sheet.title = '挂号信息'
            row_one = ['门诊编号', '主治医生', '挂号时间', '挂号科室', '状态',
                       '姓名', '身份证号', '挂号费', '社保号', '联系电话', '是否自费',
                       '性别', '年龄', '职业', '初复诊', '备注']
            for i in range(1, len(row_one) + 1):
                sheet.cell(row=1, column=i).value = row_one[i - 1]
            status_dict = {'1': '已住院', '2': '已出院', '3': '已结算', '4': '未结算', '5': '已挂号', '6': '已退号'}
            for i in record_no_list:
                max_row = sheet.max_row + 1
                try:
                    obj = Registration.objects.get(pk=i)
                except Registration.DoesNotExist:
                    return Response({'detail': '上传的参数中有无效的参数。'}, status=status.HTTP_400_BAD_REQUEST)

                if obj.is_first == 1:
                    is_first = '是'
                else:
                    is_first = '否'
                if obj.is_paying == 1:
                    is_paying = '是'
                else:
                    is_paying = '否'
                if obj.sex is True:
                    sex = '男'
                else:
                    sex = '女'
                status_info = status_dict[str(obj.status)]
                obj_info = [obj.id, obj.doctor.real_name, obj.regist_date, obj.department.department_name,
                            status_info, obj.name, obj.id_number, obj.cost, obj.social_num, obj.phone, is_paying,
                            sex, obj.age, obj.occupation, is_first, obj.remark]
                for info in range(1, len(obj_info) + 1):
                    sheet.cell(row=max_row, column=info).value = obj_info[info - 1]
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            file_name = '挂号信息%s.xls' % ctime
            file_name = urlquote(file_name)
            response['Content-Disposition'] = 'attachment; filename=%s' % file_name
            response.write(output.getvalue())
            return response
        return Response({'detail': '请传递正确的参数。'}, status=status.HTTP_400_BAD_REQUEST)


class DoctorManageViewSet(viewsets.ModelViewSet):
    """
    门诊医生管理
    """
    queryset = DoctorManage.objects.all()
    serializer_class = DoctorManageSerializer
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = (DjangoFilterBackend, OrderingFilter)
    filter_fields = ('id', 'real_name', 'department')  # 过滤查找的字段
    ordering = ('department',)
